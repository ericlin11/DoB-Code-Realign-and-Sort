import openpyxl
from openpyxl.styles import Font
import re

#2024 NYCC Code Revision -BC.xlsx

wb = openpyxl.load_workbook('2024 NYCC Code Revision - BC.xlsx')

#Range of sheets/chapters
for sheetnum in range(3, 36):
    ws = ws = wb['BC' + str(sheetnum)]

    codelist = []
    errorlist = []
    tablelink1 = ""
    tablelink2 = ""
    tablelink3 = ""
    tablelink1text = ""
    tablelink2text = ""
    tablelink3text = ""
    error = "false"

    def get_links():
        print(str(ws.cell(x, column=1)))
        try:
            global tablelink1
            global tablelink1text
            tablelink1 = ws.cell(x, column=5).hyperlink.target
            print(tablelink1)
            tablelink1text = ws.cell(x, column=5).value
            print(tablelink1text)
        except:
            print("No Table 1")
        try:
            global tablelink2
            global tablelink2text
            tablelink2 = ws.cell(x, column=6).hyperlink.target
            print(tablelink2)
            tablelink2text = ws.cell(x, column=6).value
            print(tablelink2text)
        except:
            print("No Table 2")
        try:
            global tablelink3
            global tablelink3text
            tablelink3 = ws.cell(x, column=7).hyperlink.target
            print(tablelink3)
            tablelink3text = ws.cell(x, column=7).value
            print(tablelink3text)
        except:
            print("No Table 3")


    for x in range(2, ws.max_row+1):
        if ws.cell(x, column=1).value is not None:
            nycsection = ws.cell(x,column=1).value
            nyctext = ws.cell(x,column=2).value
            get_links()
            try:
                nyc = nycsection.split(" ", 1)
                nycnumber = nyc[0]
                nyctitle = nyc[1]
                if re.search('[a-zA-Z]', nycnumber) or re.search('[@_!#$%^&*()<>?/|}{~:]', nycnumber) or re.search('[\xa0]', nycnumber):
                    print("Not a NYC Code at : " + str(ws.cell(x,column=1).value + " : " + nycsection))
                    errorlist.append({
                        "nycsection": ws.cell(x, column=1).value,
                        "nyctext": ws.cell(x, column=2).value,
                        "iccsection": ws.cell(x, column=3).value,
                        "icctext": ws.cell(x, column=4).value,
                    })
                else:
                    codelist.append({
                    "sectionnumber": nycnumber,
                    "nycsection": nycsection,
                    "nyctext": nyctext,
                    "iccsection": "",
                    "icctext": "",
                    "tablelink1": tablelink1,
                    "tablelink1text": tablelink1text,
                    "tablelink2": tablelink2,
                    "tablelink2text": tablelink2text,
                    "tablelink3": tablelink3,
                    "tablelink3text": tablelink3text,
                })
                tablelink1 = ""
                tablelink2 = ""
                tablelink3 = ""
                tablelink1text = ""
                tablelink2text = ""
                tablelink3text = ""
            except AttributeError:
                print("Not a NYC Code at : " + str(ws.cell(x,column=1)))
                errorlist.append({
                    "nycsection": ws.cell(x, column=1).value,
                    "nyctext": ws.cell(x, column=2).value,
                })
            except IndexError:
                print("Not a NYC Code at : " + str(ws.cell(x,column=1)))
                errorlist.append({
                    "nycsection": ws.cell(x, column=1).value,
                    "nyctext": ws.cell(x, column=2).value,
                    "iccsection": ws.cell(x, column=3).value,
                    "icctext": ws.cell(x, column=4).value,
                })

    for x in range(2, ws.max_row+1):
        if ws.cell(x, column=3).value is not None:
            iccsection = ws.cell(x,column=3).value
            icctext = ws.cell(x,column=4).value
            get_links()
            try:
                icc = iccsection.split(" ", 1)
                iccnumber = icc[0]
                icctitle = icc[1]
                if not any(code["sectionnumber"] == iccnumber for code in codelist):
                    print(iccnumber + " doesn't exist.")
                    codelist.append({
                                "sectionnumber": iccnumber,
                                "nycsection": "",
                                "nyctext": "",
                                "iccsection": iccsection,
                                "icctext": icctext,
                                "tablelink1": tablelink1,
                                "tablelink1text": tablelink1text,
                                "tablelink2": tablelink2,
                                "tablelink2text": tablelink2text,
                                "tablelink3": tablelink3,
                                "tablelink3text": tablelink3text,
                            })
                    tablelink1 = ""
                    tablelink2 = ""
                    tablelink3 = ""
                    tablelink1text = ""
                    tablelink2text = ""
                    tablelink3text = ""
                else:
                    for code in codelist:
                        if code["sectionnumber"] == iccnumber:
                            code["iccsection"] = iccsection
                            code["icctext"] = icctext
                            while tablelink1 not in [code["tablelink1"], code["tablelink2"], code["tablelink3"]]:
                                print("Table 1 check: " + tablelink1text + " - " + tablelink1)
                                if code["tablelink1"] is "":
                                    code["tablelink1"] = tablelink1
                                    code["tablelink1text"] = tablelink1text
                                elif code["tablelink2"] is "":
                                    code["tablelink2"] = tablelink1
                                    code["tablelink2text"] = tablelink1text
                                elif code["tablelink3"] is "":
                                    code["tablelink3"] = tablelink1
                                    code["tablelink3text"] = tablelink1text
                                break
                            while tablelink2 not in [code["tablelink1"], code["tablelink2"], code["tablelink3"]]:
                                print("Table 2 check: " + tablelink2text + " - " + tablelink2)
                                if code["tablelink1"] is "":
                                    code["tablelink1"] = tablelink2
                                    code["tablelink1text"] = tablelink2text
                                elif code["tablelink2"] is "":
                                    code["tablelink2"] = tablelink2
                                    code["tablelink2text"] = tablelink2text
                                elif code["tablelink3"] is "":
                                    code["tablelink3"] = tablelink2
                                    code["tablelink3text"] = tablelink2text
                                break
                            while tablelink3 not in [code["tablelink1"], code["tablelink2"], code["tablelink3"]]:
                                print("Table 3 check: " + tablelink3text + " - " + tablelink3)
                                if code["tablelink1"] is "":
                                    code["tablelink1"] = tablelink3
                                    code["tablelink1text"] = tablelink3text
                                elif code["tablelink2"] is "":
                                    code["tablelink2"] = tablelink3
                                    code["tablelink2text"] = tablelink3text
                                elif code["tablelink3"] is "":
                                    code["tablelink3"] = tablelink3
                                    code["tablelink3text"] = tablelink1text
                                break
                            tablelink1 = ""
                            tablelink2 = ""
                            tablelink3 = ""
                            tablelink1text = ""
                            tablelink2text = ""
                            tablelink3text = ""
                            continue
            except AttributeError:
                print("Not a ICC Code at: " + str(ws.cell(x, column=3)))
                errorlist.append({
                    "nycsection": ws.cell(x, column=1).value,
                    "nyctext": ws.cell(x, column=2).value,
                    "iccsection": ws.cell(x, column=3).value,
                    "icctext": ws.cell(x, column=4).value,
                })

    for x in errorlist:
        print("errorlist: " + str(x))
        # print("[NYC]: " + str(x["nycsection"]) + " - " + str(x["nyctext"]) + "\n " +
        #       "[ICC]: " + str(x["iccsection"]) + " - " + str(x["icctext"]) + "\n")

    codelist = sorted(codelist, key=lambda d: d["sectionnumber"])
    # codelist.sort(key=lambda s: [float(u) for u in s["sectionnumber"].split('.')])
    # for x in codelist:
    #     print(x)

    # for x in codelist:
    #     try:
    #         codelist.sort(key=lambda s: [int(u) for u in s["sectionnumber"].split('.')])
    #     except ValueError:
    #         print("Error sorting: " + str(x))
    #         break
    for x in codelist:
        print("codelist: " + str(x))
    # codelist.sort(key=lambda s: [int(u) for u in s["sectionnumber"].split('.')])

    for x in codelist:
        print(str(x["sectionnumber"]) + ": \n" +
            "[NYC]: " + str(x["nycsection"]) + " - " + str(x["nyctext"]) + "\n " +
            "[ICC]: " + str(x["iccsection"]) + " - " + str(x["icctext"]) + "\n" +
            "Table1: " + str(x["tablelink1text"]) + " - " + str(x["tablelink1"]) + "\n" +
            "Table2: " + str(x["tablelink2text"]) + " - " + str(x["tablelink2"]) + "\n" +
            "Table3: " + str(x["tablelink3text"]) + " - " + str(x["tablelink3"]) + "\n")

    startingrow = 2

    for x in errorlist:
        try:
            ws.cell(startingrow, column=1).value = x["nycsection"]
        except:
            print()
        try:
            ws.cell(startingrow, column=2).value = x["nyctext"]
        except:
            print()
        try:
            ws.cell(startingrow, column=3).value = x["iccsection"]
        except:
            print()
        try:
            ws.cell(startingrow, column=4).value = x["icctext"]
        except:
            print()
        startingrow += 1

    for x in codelist:
        print(x["sectionnumber"] + ": ")
        ws.cell(startingrow, column=1).value = x["nycsection"]
        ws.cell(startingrow, column=2).value = x["nyctext"]
        ws.cell(startingrow, column=3).value = x["iccsection"]
        ws.cell(startingrow, column=4).value = x["icctext"]
        try:
            ws.cell(startingrow, column=5).value = x["tablelink1text"]
            ws.cell(startingrow, column=5).hyperlink = x["tablelink1"]
            ws.cell(startingrow, column=5).style = "Hyperlink"
        except AttributeError:
            print("No Table 1")
        try:
            ws.cell(startingrow, column=6).value = x["tablelink2text"]
            ws.cell(startingrow, column=6).hyperlink = x["tablelink2"]
            ws.cell(startingrow, column=6).style = "Hyperlink"
        except AttributeError:
            print("No Table 2")
        try:
            ws.cell(startingrow, column=7).value = x["tablelink3text"]
            ws.cell(startingrow, column=7).hyperlink = x["tablelink3"]
            ws.cell(startingrow, column=7).style = "Hyperlink"
        except AttributeError:
            print("No Table 3")
        startingrow += 1

    for x in range(startingrow, ws.max_row+1):
        ws.cell(x, column=1).value = None
        ws.cell(x, column=2).value = None
        ws.cell(x, column=3).value = None
        ws.cell(x, column=4).value = None
        ws.cell(x, column=5).value = None
        ws.cell(x, column=6).value = None
        ws.cell(x, column=7).value = None


    print("Max row: "+ str(ws.max_row))
    print("Starting row: " + str(startingrow) +
          "\n Length of codelist: " + str(len(codelist)) +
          "\n Length of errorlist: " + str(len(errorlist)))

    if error == "true":
        ws['A1'].font = Font(color="00FFFF00")

    #Save file to sorted.xlsx
    wb.save('(sorted)2024 NYCC Code Revision - BC.xlsx')
    print("Done")